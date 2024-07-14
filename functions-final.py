from flask import Flask, request, render_template, redirect, url_for, send_file
import requests
import uuid
import time
import json
import os
import pandas as pd
import re
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
from collections import defaultdict
import matplotlib.pyplot as plt
from matplotlib import font_manager, rc
from math import cos, sin, radians
import io
import base64
import cv2
import numpy as np

app = Flask(__name__)

# 한글 폰트 설정
font_path = "C:/Windows/Fonts/malgun.ttf"  
font = font_manager.FontProperties(fname=font_path).get_name()
rc('font', family=font)

API_URL = 'https://oakceae1zv.apigw.ntruss.com/custom/v1/30641/7612742b5d4ee1285ee0cc8ca0a064584286331feacec848e0020c13ff26747a/general'
SECRET_KEY = 'SFhUSlZsaG1jVEpOelZFY294WGlHSUd6cHNDU3RLY04='


def extract_data_from_image(file):
    request_json = {
        'images': [{'format': 'jpg', 'name': 'demo'}],
        'requestId': str(uuid.uuid4()),
        'version': 'V2',
        'timestamp': int(round(time.time() * 1000))
    }
    payload = {'message': json.dumps(request_json).encode('UTF-8')}
    files = [('file', file.read())]
    headers = {'X-OCR-SECRET': SECRET_KEY}
    response = requests.post(API_URL, headers=headers, data=payload, files=files)
    response.raise_for_status()
    json_data = response.json()

    # 모든 텍스트를 하나의 큰 문자열로 결합
    string_result = ''
    for i in json_data['images'][0]['fields']:
        if i['lineBreak']:
            linebreak = '\n'
        else:
            linebreak = ' '
        string_result += i['inferText'] + linebreak
    string_result = string_result.replace(",", "")

    date_pattern = r"\d{4}-\d{2}-\d{2}"
    date_matches = re.findall(date_pattern, string_result)
    if date_matches:
        date = date_matches[0]
    else:
        date = None
    
    #영수증 패턴 읽기
    item_pattern = re.compile(r"\d{2}\*?\s+(.+?)\n\d{13}\s+(\d+)\s+(\d+)\s+(\d+)")
    matches = item_pattern.finditer(string_result)
    
    #날짜, 품목, 단가, 수량, 금액
    data = []
    for match in matches:
        item_dict = {
            "구매일자": date,
            "카테고리": match.group(0),
            "품목": match.group(1).strip(),
            "단가": match.group(2),  # 쉼표 제거
            "수량": match.group(3),
            "금액": match.group(4)  # 쉼표 제거
        }
        data.append(item_dict)
        print(f"품목: {item_dict['품목']} - 단가: {item_dict['단가']}, 수량: {item_dict['수량']}, 금액: {item_dict['금액']}")

    return data

def save_to_excel(data, filename):
    # 데이터 프레임 생성
    df = pd.DataFrame(data)
    # 파일 경로 설정
    filepath = os.path.join(r'C:\Users\soo\Desktop\soo\2024\dx\receipts', filename)
    # 데이터 프레임을 Excel 파일로 저장
    df.to_excel(filepath, index=False)
    print(f"Excel 파일이 {filepath}로 저장되었습니다.")

def add_to_excel(data):
    filepath_source = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Processed_Data.xlsx'
    filepath_destination = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Account.xlsx'  # 경로 수정
    database_filepath = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Database.xlsx'
    
    df_source = pd.read_excel(filepath_source)
    database_df = pd.read_excel(database_filepath)
    
    data_with_category = add_category_to_data(data, database_df)
    
    # 디버깅을 위한 출력문 추가
    print(f"데이터가 추가될 파일 경로: {filepath_destination}")
    print(f"추가될 데이터: {data_with_category}")

    wb = load_workbook(filepath_destination)
    ws = wb['지출리스트']
    row = ws.max_row + 1
    for item_dict in data_with_category:
        ws.cell(row=row, column=2, value=item_dict['구매일자'])
        ws.cell(row=row, column=3, value=item_dict['카테고리'])
        ws.cell(row=row, column=4, value=item_dict['품목'])
        ws.cell(row=row, column=5, value=item_dict['단가'])
        ws.cell(row=row, column=6, value=item_dict['수량'])
        ws.cell(row=row, column=7, value=item_dict['금액'])
        row += 1
    wb.save(filepath_destination)
    wb.close()
    print("데이터 전송이 완료되었습니다.")

def add_category_to_data(data, database_df):
    data_with_category = []
    #품목을 찾아 해당 카테고리명을 반환
    for item_dict in data:
        item_name = item_dict['품목']
        for col in database_df.columns[1:]:
            if item_name in database_df[col].values:
                item_dict['카테고리'] = database_df.iloc[:, 0][database_df[col] == item_name].values[0]
                break
        else:
            item_dict['카테고리'] = None
        data_with_category.append(item_dict)
    return data_with_category

def calculate_purchase_cycle(purchase_history):
    # 분류별 구매 데이터 정리
    items = defaultdict(list)
    for purchase in purchase_history:
        date = purchase["날짜"]
        items[purchase["분류"]].append({"날짜": date, "수량": purchase["수량"]})

    # 주기 계산
    item_cycles = {}
    for 분류, purchases in items.items():
        if len(purchases) < 2:
            # 구매 기록이 하나뿐인 경우 주기를 계산할 수 없음
            item_cycles[분류] = 0
            continue

        purchases.sort(key=lambda x: x["날짜"])  # 날짜순 정렬
        total_quantity = sum(p["수량"] for p in purchases[:-1])  # 마지막 구매 수량 제외
        weighted_sum = 0

        for i in range(1, len(purchases)):
            days_between = (purchases[i]["날짜"] - purchases[i-1]["날짜"]).days
            weighted_sum += days_between * purchases[i-1]["수량"]

        if total_quantity > 0:  # 분모가 0이 되는 경우를 방지
            purchase_cycle = weighted_sum / total_quantity
        else:
            purchase_cycle = 0  # 혹시나 모든 수량이 마지막 구매에 해당하면 0으로 설정

        item_cycles[분류] = purchase_cycle

    return item_cycles

def group_cycles(sorted_cycles, max_diff=7):
    groups = []
    group = []

    for i in range(len(sorted_cycles)):
        if sorted_cycles[i][1] == 0:
            continue
        if not group or sorted_cycles[i][1] - group[0][1] <= max_diff:
            group.append(sorted_cycles[i])
        else:
            groups.append(group)
            group = [sorted_cycles[i]]

    if group:
        groups.append(group)

    # 그룹 내에서 가장 큰 값과 가장 작은 값의 차이가 7 이하인지 확인
    final_groups = []
    for group in groups:
        if max(group, key=lambda x: x[1])[1] - min(group, key=lambda x: x[1])[1] <= max_diff:
            final_groups.append(group)

    return final_groups

def calculate_group_averages(groups):
    averages = {}
    for group in groups:
        group_name = ", ".join([item[0] for item in group])
        group_average = sum(item[1] for item in group) / len(group)
        averages[group_name] = group_average
    return averages

def get_last_purchase_dates(groups, purchase_history):
    last_purchase_dates = {}
    for group in groups:
        group_name = ", ".join([item[0] for item in group])
        last_dates = []
        for item in group:
            for purchase in purchase_history:
                if purchase["분류"] == item[0]:
                    last_dates.append(purchase["날짜"])
        last_purchase_dates[group_name] = max(last_dates) if last_dates else None
    return last_purchase_dates

def recommend_next_purchase_dates(averages, last_purchase_dates, latest_date, num_recommendations=1):
    recommendations = []
    for group_name, average in averages.items():
        if last_purchase_dates[group_name]:
            next_purchase_date = last_purchase_dates[group_name] + timedelta(days=round(average))
            while next_purchase_date <= latest_date:
                next_purchase_date += timedelta(days=round(average))
            for _ in range(num_recommendations):
                recommendations.append(f"{next_purchase_date.strftime('%Y년 %m월 %d일')}에 {group_name}를 사는 건 어떨까요?")
                next_purchase_date += timedelta(days=round(average))
    return recommendations

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import font_manager, rc
from datetime import timedelta
import numpy as np
import io
import base64
from math import cos, sin, radians

def generate_chart(file_path='C:/Users/soo/Desktop/soo/2024/dx/receipts/Account.xlsx'):
    # 엑셀 파일에서 데이터 읽기
    df = pd.read_excel(file_path, sheet_name='지출리스트', header=5, usecols='B:G')

    # 최근 3개월 동안의 데이터 필터링
    df['날짜'] = pd.to_datetime(df['날짜'], errors='coerce')

    latest_date = df['날짜'].max()
    three_months_ago = latest_date - timedelta(days=90)
    filtered_df = df[df['날짜'] >= three_months_ago]

    # 분류별 금액 합산
    category_sum = filtered_df.groupby('분류')['금액'].sum()

    # 도넛 차트 색상 설정
    colors = ['#ff9999','#66b3ff','#99ff99','#ffcc99','#c2c2f0','#ffb3e6', '#c4e17f', '#76d7c4']

    # 도넛 차트 생성
    fig, ax = plt.subplots(figsize=(10, 8))  # 크기 조정
    wedges, texts, autotexts = ax.pie(
        category_sum,
        labels=category_sum.index,
        startangle=90,
        wedgeprops={'width': 0.4, 'edgecolor': 'w'},  # 도넛 차트 스타일 조정
        autopct='%1.1f%%',
        pctdistance=0.85,  # 퍼센트 위치 조정
        colors=colors  # 색상 추가
    )

    # 차트 중앙에 원 그리기 (도넛 모양 완성)
    centre_circle = plt.Circle((0, 0), 0.70, fc='white')
    fig.gca().add_artist(centre_circle)

    # 범례 설정
    plt.legend(
        loc='upper left',
        bbox_to_anchor=(1, 1),
        title="Categories"
    )

    # 이미지 저장을 위한 바이트 스트림 설정
    img = io.BytesIO()
    plt.savefig(img, format='png', bbox_inches='tight')  # 여백 문제 해결
    img.seek(0)
    chart_data = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()

    # 수량을 기준으로 상위 5개 품목 계산
    top_items = filtered_df.groupby('품목')['수량'].sum().nlargest(5)
    top_items_list = [f"{item}: {int(quantity)}개" for item, quantity in top_items.items()]

    return chart_data, top_items_list

from flask import Flask, send_file
from openpyxl.styles import Font, PatternFill

@app.route('/')
def index():
    chart_data, top_items_list = generate_chart()
    
    # '지출리스트' 시트에서 데이터를 읽고 분석하여 구매 주기와 추천 날짜를 계산
    filepath = 'C:/Users/soo/Desktop/soo/2024/dx/receipts/Account.xlsx'
    data = pd.read_excel(filepath, sheet_name='지출리스트', header=5, usecols='B,C,F')
    data['날짜'] = pd.to_datetime(data['날짜'], errors='coerce')
    cleaned_data = data.dropna()
    purchase_history = [
        {"날짜": row['날짜'], "분류": row['분류'], "수량": row['수량']}
        for _, row in cleaned_data.iterrows()
    ]
    item_cycles = calculate_purchase_cycle(purchase_history)
    sorted_cycles = sorted(item_cycles.items(), key=lambda x: x[1])
    groups = group_cycles(sorted_cycles)
    averages = calculate_group_averages(groups)
    last_purchase_dates = get_last_purchase_dates(groups, purchase_history)
    latest_date = cleaned_data['날짜'].max()
    recommendations = recommend_next_purchase_dates(averages, last_purchase_dates, latest_date)

    return render_template('upload.html', chart_data=chart_data, top_items_list=top_items_list, recommendations=recommendations)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        data = extract_data_from_image(file)
        save_to_excel(data, 'Processed_Data.xlsx')
        add_to_excel(data)
        return redirect(url_for('index'))
    return redirect(url_for('index'))

@app.route('/download_excel')
def download_excel():
    # 원본 파일 경로
    filepath = r'C:/Users/soo/Desktop/soo/2024/dx/receipts/Account.xlsx'
    # 수정된 파일을 저장할 임시 경로
    temp_filepath = r'C:/Users/soo/Desktop/soo/2024/dx/receipts/가계부.xlsx'

    # Excel 파일 읽기 및 수정
    try:
        # 데이터프레임으로 엑셀 데이터 읽기
        df = pd.read_excel(filepath, sheet_name='지출리스트')
        
        # 7~13행 삭제
        modified_df = df.drop(df.index[5:13])
        modified_df.reset_index(drop=True, inplace=True)
        
        # 변경된 데이터프레임을 엑셀 파일에 저장
        with pd.ExcelWriter(temp_filepath, engine='openpyxl') as writer:
            modified_df.to_excel(writer, index=False, sheet_name='지출리스트')
        
        # openpyxl로 엑셀 파일 다시 열기
        wb = load_workbook(temp_filepath)
        ws = wb['지출리스트']
        
        # B6부터 G6까지 셀에 스타일 적용
        bold_font = Font(bold=True)
        fill_color = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
        for row in ws['B6:G6']:
            for cell in row:
                cell.font = bold_font
                cell.fill = fill_color

        
        # 스타일 적용된 파일 저장
        wb.save(temp_filepath)

        # 수정된 파일을 사용자에게 제공
        return send_file(temp_filepath, as_attachment=True)

    except Exception as e:
        return str(e)  # 예외 발생 시 에러 메시지 반환

if __name__ == '__main__':
    app.run(debug=True)