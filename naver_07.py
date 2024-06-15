from flask import Flask, request, render_template, redirect, url_for
import requests
import uuid
import time
import json
import os
import pandas as pd
import re
from openpyxl import load_workbook, Workbook
import datetime
from collections import defaultdict

app = Flask(__name__)

API_URL = 'https://oakceae1zv.apigw.ntruss.com/custom/v1/30641/7612742b5d4ee1285ee0cc8ca0a064584286331feacec848e0020c13ff26747a/general'
SECRET_KEY = 'SFhUSlZsaG1jVEpOelZFY294WGlHSUd6cHNDU3RLY04='

def extract_data_from_image(file):
    request_json = {
        'images': [{'format': 'jpg', 'name': 'demo'}],
        'requestId': str(uuid.uuid4()),
        'version': 'V2',
        'timestamp': int(round(time.time() * 1000))
    }
    payload = {'message': json.dumps(request_json).encode('UTF-8')}
    files = [('file', file.read())]
    headers = {'X-OCR-SECRET': SECRET_KEY}
    response = requests.post(API_URL, headers=headers, data=payload, files=files)
    response.raise_for_status()
    json_data = response.json()

    # 모든 텍스트를 하나의 큰 문자열로 결합
    string_result = ''
    for i in json_data['images'][0]['fields']:
        if i['lineBreak']:
            linebreak = '\n'
        else:
            linebreak = ' '
        string_result += i['inferText'] + linebreak
    string_result = string_result.replace(",", "")

    date_pattern = r"\d{4}-\d{2}-\d{2}"
    date_matches = re.findall(date_pattern, string_result)
    if date_matches:
        date = date_matches[0]
    else:
        date = None

    # 하나의 유연한 패턴을 사용하여 상품 정보 추출
    item_pattern = re.compile(r"(\d{2}\*?\s+.+?)\s+(\d+)\s+(\d+)\s+(\d+)")
    matches = item_pattern.finditer(string_result)

    data = []
    for match in matches:
        try:
            item_dict = {
                "구매일자": date,
                "카테고리": match.group(1).strip().replace('\n', ' '),
                "품목": match.group(1).strip().replace('\n', ' '),
                "단가": int(match.group(2)),
                "수량": int(match.group(3)),
                "금액": int(match.group(4))
            }
            data.append(item_dict)
            print(f"품목: {item_dict['품목']} - 단가: {item_dict['단가']}, 수량: {item_dict['수량']}, 금액: {item_dict['금액']}")
        except IndexError:
            print("매칭되지 않는 항목을 건너뜁니다.")
            continue

    return data

def save_to_excel(data, filename):
    # 데이터 프레임 생성
    df = pd.DataFrame(data)
    # 파일 경로 설정
    filepath = os.path.join(r'C:\Users\soo\Desktop\soo\2024\dx\receipts', filename)
    # 데이터 프레임을 Excel 파일로 저장
    df.to_excel(filepath, index=False)
    print(f"Excel 파일이 {filepath}로 저장되었습니다.")

def add_to_excel(data):
    filepath_source = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Processed_Data.xlsx'
    filepath_destination = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Account.xlsx'
    database_filepath = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Database.xlsx'
    
    df_source = pd.read_excel(filepath_source)
    database_df = pd.read_excel(database_filepath)
    
    data_with_category = add_category_to_data(data, database_df)
    
    wb = load_workbook(filepath_destination)
    ws = wb['지출리스트']
    row = ws.max_row + 1

    # 데이터 추가 과정 디버깅
    for item_dict in data_with_category:
        print(f"추가할 데이터: {item_dict}")
        ws.cell(row=row, column=2, value=item_dict['구매일자'])
        ws.cell(row=row, column=3, value=item_dict['카테고리'])
        ws.cell(row=row, column=4, value=item_dict['품목'])
        ws.cell(row=row, column=5, value=item_dict['단가'])
        ws.cell(row=row, column=6, value=item_dict['수량'])
        ws.cell(row=row, column=7, value=item_dict['금액'])
        row += 1

    wb.save(filepath_destination)
    wb.close()
    print("데이터 전송이 완료되었습니다.")

def add_category_to_data(data, database_df):
    data_with_category = []
    for item_dict in data:
        item_name = item_dict['품목']
        for col in database_df.columns[1:]:
            if item_name in database_df[col].values:
                item_dict['카테고리'] = database_df.iloc[:, 0][database_df[col] == item_name].values[0]
                break
        else:
            item_dict['카테고리'] = None
        data_with_category.append(item_dict)
    return data_with_category

# 구매 주기 계산 함수
def calculate_purchase_cycle(purchase_history):
    # 분류별 구매 데이터 정리
    items = defaultdict(list)
    for purchase in purchase_history:
        date = purchase["날짜"]
        items[purchase["분류"]].append({"날짜": date, "수량": purchase["수량"]})

    # 주기 계산
    item_cycles = {}
    for 분류, purchases in items.items():
        if len(purchases) < 2:
            # 구매 기록이 하나뿐인 경우 주기를 계산할 수 없음
            item_cycles[분류] = 0
            continue

        purchases.sort(key=lambda x: x["날짜"])  # 날짜순 정렬
        total_quantity = sum(p["수량"] for p in purchases[:-1])  # 마지막 구매 수량 제외
        weighted_sum = 0

        for i in range(1, len(purchases)):
            days_between = (purchases[i]["날짜"] - purchases[i-1]["날짜"]).days
            weighted_sum += days_between * purchases[i-1]["수량"]

        if total_quantity > 0:  # 분모가 0이 되는 경우를 방지
            purchase_cycle = weighted_sum / total_quantity
        else:
            purchase_cycle = 0  # 혹시나 모든 수량이 마지막 구매에 해당하면 0으로 설정

        item_cycles[분류] = purchase_cycle

    return item_cycles

# 각 구매 주기 간 차이가 7 이하인 분류들을 그룹화
def group_cycles(sorted_cycles, max_diff=7):
    groups = []
    group = []

    for i in range(len(sorted_cycles)):
        if sorted_cycles[i][1] == 0:
            continue
        if not group or sorted_cycles[i][1] - group[0][1] <= max_diff:
            group.append(sorted_cycles[i])
        else:
            groups.append(group)
            group = [sorted_cycles[i]]

    if group:
        groups.append(group)

    # 그룹 내에서 가장 큰 값과 가장 작은 값의 차이가 7 이하인지 확인
    final_groups = []
    for group in groups:
        if max(group, key=lambda x: x[1])[1] - min(group, key=lambda x: x[1])[1] <= max_diff:
            final_groups.append(group)

    return final_groups

# 각 그룹의 단순 평균 계산
def calculate_group_averages(groups):
    averages = {}
    for group in groups:
        group_name = ", ".join([item[0] for item in group])
        group_average = sum(item[1] for item in group) / len(group)
        averages[group_name] = group_average
    return averages

# 그룹별 마지막 구매 일자 계산
def get_last_purchase_dates(groups, purchase_history):
    last_purchase_dates = {}
    for group in groups:
        group_name = ", ".join([item[0] for item in group])
        last_dates = []
        for item in group:
            for purchase in purchase_history:
                if purchase["분류"] == item[0]:
                    last_dates.append(purchase["날짜"])
        last_purchase_dates[group_name] = max(last_dates) if last_dates else None
    return last_purchase_dates

# 그룹별 다음 구매 일자 도출 함수
def recommend_next_purchase_dates(averages, last_purchase_dates, latest_date, num_recommendations=1):
    recommendations = []
    for group_name, average in averages.items():
        if last_purchase_dates[group_name]:
            next_purchase_date = last_purchase_dates[group_name] + datetime.timedelta(days=round(average))
            while next_purchase_date <= latest_date:
                next_purchase_date += datetime.timedelta(days=round(average))
            for _ in range(num_recommendations):
                recommendations.append(f"{next_purchase_date.strftime('%Y년 %m월 %d일')}은 {group_name} 사는 날입니다")
                next_purchase_date += datetime.timedelta(days=round(average))
    return recommendations

@app.route('/')
def index():
    # '지출리스트' 시트에서 데이터를 읽고 분석하여 구매 주기와 추천 날짜를 계산
    filepath = 'C:/Users/soo/Desktop/soo/2024/dx/receipts/Account.xlsx'
    data = pd.read_excel(filepath, sheet_name='지출리스트', header=5, usecols='B,C,F')
    data['날짜'] = pd.to_datetime(data['날짜'], errors='coerce')
    cleaned_data = data.dropna()
    purchase_history = [
        {"날짜": row['날짜'], "분류": row['분류'], "수량": row['수량']}
        for _, row in cleaned_data.iterrows()
    ]
    item_cycles = calculate_purchase_cycle(purchase_history)
    sorted_cycles = sorted(item_cycles.items(), key=lambda x: x[1])
    groups = group_cycles(sorted_cycles)
    averages = calculate_group_averages(groups)
    last_purchase_dates = get_last_purchase_dates(groups, purchase_history)
    latest_date = cleaned_data['날짜'].max()
    recommendations = recommend_next_purchase_dates(averages, last_purchase_dates, latest_date)

    return render_template('upload.html', recommendations=recommendations)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        data = extract_data_from_image(file)
        save_to_excel(data, 'C:/Users/soo/Desktop/soo/2024/dx/receipts/Processed_Data.xlsx')
        add_to_excel(data)
        return render_template('result.html', data=data)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
