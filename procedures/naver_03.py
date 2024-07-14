from flask import Flask, request, render_template, redirect, url_for
import requests
import uuid
import time
import json
import os
import pandas as pd
import re
from openpyxl import load_workbook

app = Flask(__name__)

API_URL = 'https://oakceae1zv.apigw.ntruss.com/custom/v1/30641/7612742b5d4ee1285ee0cc8ca0a064584286331feacec848e0020c13ff26747a/general'
SECRET_KEY = 'SFhUSlZsaG1jVEpOelZFY294WGlHSUd6cHNDU3RLY04='

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        data = process_image(file)
        transfer_data_to_account_excel('Processed_Data.xlsx', 'Account.xlsx')
        return render_template('result.html', data=data)
    return redirect(url_for('index'))

def save_to_excel(data, filename):
    df = pd.DataFrame(data)
    filepath = os.path.join(r'C:\Users\soo\Desktop\soo\2024\dx\receipts', filename)
    df.to_excel(filepath, index=False)
    print(f"Excel 파일이 {filepath}로 저장되었습니다.")

def process_image(file):
    request_json = {
        'images': [{'format': 'jpg', 'name': 'demo'}],
        'requestId': str(uuid.uuid4()),
        'version': 'V2',
        'timestamp': int(round(time.time() * 1000))
    }
    payload = {'message': json.dumps(request_json).encode('UTF-8')}
    files = [('file', file.read())]
    headers = {'X-OCR-SECRET': SECRET_KEY}
    response = requests.post(API_URL, headers=headers, data=payload, files=files)
    response.raise_for_status()
    json_data = response.json()

         # 모든 텍스트를 하나의 큰 문자열로 결합
    string_result = ''
    for i in json_data['images'][0]['fields']:
        if i['lineBreak']:
            linebreak = '\n'
        else:
            linebreak = ' '
        string_result += i['inferText'] + linebreak
    string_result = string_result.replace(",", "")

    pattern = re.compile(r"\d{2}\*?\s+(.+?)\n\d{13}\s+(\d+)\s+(\d+)\s+(\d+)")
    matches = pattern.finditer(string_result)

    data = []
    for match in matches:
        item_dict = {
            "품목": match.group(1).strip(),
            "단가": match.group(2),  # 쉼표 제거
            "수량": match.group(3),
            "금액": match.group(4)  # 쉼표 제거
        }
        data.append(item_dict)
        print(f"품목: {item_dict['품목']} - 단가: {item_dict['단가']}, 수량: {item_dict['수량']}, 금액: {item_dict['금액']}")

    save_to_excel(data, 'Processed_Data.xlsx')
    return data

def transfer_data_to_account_excel(source_file_name, destination_file_name):
    source_path = os.path.join(r'C:\Users\soo\Desktop\soo\2024\dx\receipts', source_file_name)
    destination_path = os.path.join(r'C:\Users\soo\Desktop\soo\2024\dx\receipts', destination_file_name)
    df_source = pd.read_excel(source_path)
    wb = load_workbook(destination_path)
    ws = wb['지출리스트']
    row = ws.max_row + 1
    for index, row_data in df_source.iterrows():
        ws.cell(row=row, column=4, value=row_data['품목'])
        ws.cell(row=row, column=5, value=row_data['단가'])
        ws.cell(row=row, column=6, value=row_data['수량'])
        ws.cell(row=row, column=7, value=row_data['금액'])
        row += 1
    wb.save(destination_path)
    wb.close()
    print("데이터가 Account.xlsx 파일로 전송되었습니다.")

if __name__ == '__main__':
    app.run(debug=True)