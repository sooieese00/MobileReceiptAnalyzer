from flask import Flask, request, render_template, redirect, url_for
import requests
import uuid
import time
import json
import os
import pandas as pd
import re
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

API_URL = 'https://oakceae1zv.apigw.ntruss.com/custom/v1/30641/7612742b5d4ee1285ee0cc8ca0a064584286331feacec848e0020c13ff26747a/general'
SECRET_KEY = 'SFhUSlZsaG1jVEpOelZFY294WGlHSUd6cHNDU3RLY04='

def extract_data_from_image(file):
    request_json = {
        'images': [{'format': 'jpg', 'name': 'demo'}],
        'requestId': str(uuid.uuid4()),
        'version': 'V2',
        'timestamp': int(round(time.time() * 1000))
    }
    payload = {'message': json.dumps(request_json).encode('UTF-8')}
    files = [('file', file.read())]
    headers = {'X-OCR-SECRET': SECRET_KEY}
    response = requests.post(API_URL, headers=headers, data=payload, files=files)
    response.raise_for_status()
    json_data = response.json()

    # 모든 텍스트를 하나의 큰 문자열로 결합
    string_result = ''
    for i in json_data['images'][0]['fields']:
        if i['lineBreak']:
            linebreak = '\n'
        else:
            linebreak = ' '
        string_result += i['inferText'] + linebreak
    string_result = string_result.replace(",", "")

    date_pattern = r"\d{4}-\d{2}-\d{2}"
    date_matches = re.findall(date_pattern, string_result)
    if date_matches:
        date = date_matches[0]
    else:
        date = None
    
    item_pattern = re.compile(r"\d{2}\*?\s+(.+?)\n\d{13}\s+(\d+)\s+(\d+)\s+(\d+)")
    matches = item_pattern.finditer(string_result)

    data = []
    for match in matches:
        item_dict = {
            "구매일자": date,
            "카테고리": match.group(0),
            "품목": match.group(1).strip(),
            "단가": match.group(2),  # 쉼표 제거
            "수량": match.group(3),
            "금액": match.group(4)  # 쉼표 제거
        }
        data.append(item_dict)
        print(f"품목: {item_dict['품목']} - 단가: {item_dict['단가']}, 수량: {item_dict['수량']}, 금액: {item_dict['금액']}")

    return data

def save_to_excel(data, filename):
    # 데이터 프레임 생성
    df = pd.DataFrame(data)
    # 파일 경로 설정
    filepath = os.path.join(r'C:\Users\soo\Desktop\soo\2024\dx\receipts', filename, filename)
    # 데이터 프레임을 Excel 파일로 저장
    df.to_excel(filepath, index=False)
    print(f"Excel 파일이 {filepath}로 저장되었습니다.")

def add_to_excel(data):
    filepath_source = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Processed_Data.xlsx'
    filepath_destination = 'C:/Users/soo/Desktop/soo/2024/dx/receipt/Account.xlsx'
    database_filepath = r'C:\Users\soo\Desktop\soo\2024\dx\receipts\Database.xlsx'
    
    df_source = pd.read_excel(filepath_source)
    database_df = pd.read_excel(database_filepath)
    
    data_with_category = add_category_to_data(data, database_df)
    
    wb = load_workbook(filepath_destination)
    ws = wb['지출리스트']
    row = ws.max_row + 1
    for item_dict in data_with_category:
        ws.cell(row=row, column=2, value=item_dict['구매일자'])
        ws.cell(row=row, column=3, value=item_dict['카테고리'])
        ws.cell(row=row, column=4, value=item_dict['품목'])
        ws.cell(row=row, column=5, value=item_dict['단가'])
        ws.cell(row=row, column=6, value=item_dict['수량'])
        ws.cell(row=row, column=7, value=item_dict['금액'])
        row += 1
    wb.save(filepath_destination)
    wb.close()
    print("데이터 전송이 완료되었습니다.")

def add_category_to_data(data, database_df):
    data_with_category = []
    for item_dict in data:
        item_name = item_dict['품목']
        for col in database_df.columns[1:]:
            if item_name in database_df[col].values:
                item_dict['카테고리'] = database_df.iloc[:, 0][database_df[col] == item_name].values[0]
                break
        else:
            item_dict['카테고리'] = None
        data_with_category.append(item_dict)
    return data_with_category



@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        data = extract_data_from_image(file)
        save_to_excel(data, 'C:/Users/soo/Desktop/soo/2024/dx/receipts/Processed_Data.xlsx')
        add_to_excel(data)
        return render_template('result.html', data=data)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)