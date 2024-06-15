import pandas as pd
from openpyxl import load_workbook

# 정확한 파일 경로를 입력하세요
source_file = 'C:/Users/soo/Desktop/soo/2024/dx/receipts/Processed_Data.xlsx'
destination_file = 'C:/Users/soo/Desktop/soo/2024/dx/receipts/Account.xlsx'

# 소스 파일에서 데이터를 읽어옵니다
df_source = pd.read_excel(source_file)

# 목표 파일을 로드합니다
wb = load_workbook(destination_file)
# 정확한 시트 이름을 입력하세요
ws = wb['지출리스트']  # 시트 이름이 맞는지 확인하세요

# 첫 번째 빈 행을 찾습니다
row = ws.max_row + 1

# DataFrame의 각 행에 대해 반복하며 데이터를 시트에 작성합니다
for index, row_data in df_source.iterrows():
    ws.cell(row=row, column=4, value=row_data['품목'])
    ws.cell(row=row, column=5, value=row_data['단가'])
    ws.cell(row=row, column=6, value=row_data['수량'])
    ws.cell(row=row, column=7, value=row_data['금액'])
    row += 1

# 변경 사항을 저장하고 파일을 닫습니다
wb.save(destination_file)
wb.close()

print("데이터 전송이 완료되었습니다.")